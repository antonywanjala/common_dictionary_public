# -*- coding: utf-8 -*-
"""
UNIVERSAL RAW BINARY-FIRST CARTESIAN PROGRAMMATIC ENGINE (V6.4)
- Exhaustively evaluates all 2^n permutations for allocated bit lengths.
- Prevents OOM (Out of Memory) crashes via lazy evaluation generators.
- Isolates individual variable binary contributions in previews and CSV exports.
- Dual-streams to CSV and XLSX. Auto-chunks XLSX sheets at 1,048,500 rows.
- Isolates output documentation into dynamic, epoch-stamped session folders.
*Upgrade: Supports ingestion of CSV/XLSX binary sets for incremental n-length combinatorics mapping.*
"""

import os
import sys
import types
import csv
import itertools
import math
import time

# Attempt to load openpyxl for Excel output/input, falling back gracefully if missing
try:
    from openpyxl import Workbook, load_workbook

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class BiDict(dict):
    """Native Bidirectional Dictionary tracking object namespace relationships."""

    def __init__(self, *args, **kwargs):
        super(BiDict, self).__init__(*args, **kwargs)
        self.inverse = {}
        for key, value in self.items():
            self.inverse[value] = key

    def __setitem__(self, key, value):
        super(BiDict, self).__setitem__(key, value)
        self.inverse[value] = key

    def get_key_by_value(self, value):
        return self.inverse.get(value)


def load_sequence_file(filepath):
    """Parses binary strings from provided CSV or XLSX matrices."""
    sequences = []
    try:
        if filepath.endswith('.csv'):
            with open(filepath, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                for row in reader:
                    sequences.extend([str(x).strip() for x in row if str(x).strip()])
        elif filepath.endswith('.xlsx') and HAS_OPENPYXL:
            wb = load_workbook(filepath, read_only=True, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                sequences.extend([str(x).strip() for x in row if x is not None and str(x).strip()])
        else:
            print(f"[ERROR] Unsupported file format or openpyxl missing for: {filepath}")
    except Exception as e:
        print(f"[ERROR] Matrix file compilation failure: {e}")

    # Return distinct unique operational strings
    return sorted(list(set(sequences)))


class BinarySequenceNgramEngine:
    """Modeled after Script2 SetTransitionEngine: Generates n-length sequence permutations/combinations."""

    def __init__(self, base_sequences, start_length, max_length, gen_method='hybrid'):
        self.base_sequences = sorted(list(set(base_sequences)))
        self.start_length = start_length
        self.max_length = max_length
        self.gen_method = gen_method
        self.structural_matrix = {}

    def generate_and_map(self):
        """Builds programmatic state sequence sets iteratively from n to n+k."""
        print(f"\n[INFO] Initializing Sequence N-Gram Generation Engine (Method: {self.gen_method.upper()})")

        for length in range(self.start_length, self.max_length + 1):
            start_ns = time.perf_counter_ns()

            if self.gen_method == 'combinations':
                raw_states = list(itertools.combinations_with_replacement(self.base_sequences, length))
            elif self.gen_method == 'permutations_strict':
                raw_states = list(itertools.permutations(self.base_sequences, length))
            else:
                raw_states = list(itertools.product(self.base_sequences, repeat=length))

            processed_states = []
            for x in raw_states:
                state_list = list(x)
                if self.gen_method == 'combinations':
                    state_list.sort()
                processed_states.append(tuple(state_list))

            unique_states = sorted(list(set(processed_states)))
            self.structural_matrix[length] = unique_states

            compute_ms = (time.perf_counter_ns() - start_ns) / 1_000_000
            print(
                f"  -> Sequence Length [n={length}]: Compiled {len(unique_states):,} unique programmatic states ({compute_ms:.2f}ms).")


class CartesianBitContinuum:
    """Handles exhaustive low-level dynamic memory permutations."""

    def __init__(self, name, var_type, base_value, allocated_bits):
        self.name = name
        self.type = var_type
        self.base_value = base_value
        self.allocated_bits = allocated_bits
        # Total permutations is exactly 2^n
        self.total_permutations = 1 << self.allocated_bits

    def decode_binary(self, bit_str):
        """Decodes pure Cartesian raw bit configurations into programmatic values."""
        if len(bit_str) < self.allocated_bits:
            return "ERR (Truncated)"

        if self.type == "NUMERIC":
            return int(bit_str, 2)
        else:
            chars = []
            for i in range(0, len(bit_str), 8):
                byte = bit_str[i:i + 8]
                if len(byte) < 8: break
                chars.append(chr(int(byte, 2)))

            # Use repr() to safely escape unprintable ASCII generated during brute force
            decoded_str = "".join(chars)
            return repr(decoded_str)


class CartesianCorePipeline:
    def __init__(self):
        # Capture the epoch timestamp at the exact moment of session invocation
        self.session_epoch = int(time.time())
        self.session_folder_name = f"Session_{self.session_epoch}"

        self.continuum_map = BiDict()
        self.staging_metadata = {}
        self.variable_order = []
        self.wattage = 65.0
        self.total_matrix_size = 0

    def execute_introspection(self, file_path):
        if not os.path.exists(file_path):
            print(f"[ERROR] Target path invalid: {file_path}")
            return False

        self.continuum_map.clear()
        self.staging_metadata.clear()
        self.variable_order.clear()

        print(f"\n[PHASE 1] Dynamic Introspection Pipeline executing at: {file_path}")

        dir_name = os.path.dirname(os.path.abspath(file_path))
        sys.path.insert(0, dir_name)
        original_cwd = os.getcwd()
        os.chdir(dir_name)

        try:
            namespace = {"__file__": os.path.abspath(file_path), "__name__": "__main__"}
            with open(file_path, "r", encoding="utf-8") as f:
                exec(f.read(), namespace)

            for obj_name, obj_instance in namespace.items():
                if obj_name.startswith("__") or isinstance(obj_instance, (types.ModuleType, types.FunctionType, type)):
                    continue

                if hasattr(obj_instance, "__dict__") and not isinstance(obj_instance, dict):
                    print(f"  [+] Introspecting Object Instance Structure: '{obj_name}'")
                    for attr_name, attr_val in obj_instance.__dict__.items():
                        if attr_name.startswith("__"): continue
                        composite_name = f"{obj_name}.{attr_name}"
                        self.variable_order.append(composite_name)

                        var_type = "NUMERIC" if isinstance(attr_val, (int, float)) and not isinstance(attr_val,
                                                                                                      bool) else "LITERAL"
                        self.staging_metadata[composite_name] = {"val": attr_val, "type": var_type}
                        print(f"    [Captured {var_type}] '{composite_name}': Base Value -> {attr_val}")
                else:
                    self.variable_order.append(obj_name)
                    var_type = "NUMERIC" if isinstance(obj_instance, (int, float)) and not isinstance(obj_instance,
                                                                                                      bool) else "LITERAL"
                    self.staging_metadata[obj_name] = {"val": obj_instance, "type": var_type}
                    print(f"    [Captured {var_type}] '{obj_name}': Base Value -> {obj_instance}")

            return len(self.variable_order) > 0
        except Exception as e:
            print(f"[ERROR] Introspection pipeline failure: {e}")
            return False
        finally:
            os.chdir(original_cwd)

    def configure_memory_footprints(self):
        print("\n[PHASE 2] Binary Sequence Parameter Configuration")
        print("Define the exact exhaustive binary boundaries for each captured variable:")

        for var in self.variable_order:
            meta = self.staging_metadata[var]
            default_bits = 4 if meta["type"] == "NUMERIC" else 16

            try:
                user_bits = input(
                    f"  -> Enter bit-length for '{var}' (Type: {meta['type']}, Default: {default_bits}): ").strip()
                allocated_bits = int(user_bits) if user_bits else default_bits
            except ValueError:
                allocated_bits = default_bits

            continuum_obj = CartesianBitContinuum(var, meta["type"], meta["val"], allocated_bits)
            self.continuum_map[var] = continuum_obj

        print("\n[PHASE 3] Computing exhaustive Cartesian logic states from raw bit-blueprints...")

        self.total_matrix_size = math.prod(self.continuum_map[v].total_permutations for v in self.variable_order)
        print(f"--> Success. Total Exhaustive Generated Binary Combinations: {self.total_matrix_size:,}")
        if self.total_matrix_size > 1048500:
            print(
                "  [!] MASSIVE MATRIX DETECTED: XLSX generation will auto-chunk across multiple sheets to bypass the 1,048,576 row limit.")

    def get_cartesian_generator(self):
        """Returns a lazy-evaluating generator of all permutations to prevent RAM exhaustion."""
        spaces = [range(self.continuum_map[v].total_permutations) for v in self.variable_order]
        return itertools.product(*spaces)

    def extract_state_details(self, combo_tuple):
        """Processes a raw coordinate tuple into binary strings and decoded values."""
        binary_contributions = {}
        decoded_values = {}
        full_seq = ""

        for idx, var in enumerate(self.variable_order):
            continuum = self.continuum_map[var]
            bin_str = format(combo_tuple[idx], f"0{continuum.allocated_bits}b")

            binary_contributions[var] = bin_str
            full_seq += bin_str
            decoded_values[var] = continuum.decode_binary(bin_str)

        return full_seq, binary_contributions, decoded_values


# ==========================================
# EXECUTION ORCHESTRATOR
# ==========================================
if __name__ == "__main__":
    pipeline = CartesianCorePipeline()

    print("=================================================================")
    print("   UNIVERSAL RAW BINARY-FIRST CARTESIAN PROGRAMMATIC ENGINE      ")
    print("=================================================================")

    try:
        watt_in = input("Enter hardware configuration wattage (e.g., 65.0): ").strip()
        pipeline.wattage = float(watt_in) if watt_in else 65.0
    except ValueError:
        pipeline.wattage = 65.0

    path_input = input(
        "Enter target absolute script file path (absfilepath) (Leave blank to skip introspection): ").strip().strip('"')

    if path_input and pipeline.execute_introspection(path_input):
        pipeline.configure_memory_footprints()

    while True:
        print("\n------------------------------------------------------------")
        print("Select State Synthesis Mode:")
        print("1) Preview Generated Binary Continuum Sequences & Decoded Value States")
        print("2) Export Solution Set Traversals with Resource Profiling to CSV & XLSX")
        print("3) Input Concatenated Binary Sequence -> Rebuild System State")
        print("4) [Functionality1] Matrix Import: Binary Sequence N-Gram Combinatorics Engine")
        print("5) Exit Unified Core Pipeline")
        choice = input("Selection (1-5): ").strip()

        if choice == "1":
            print("\n[PREVIEW MODE] Displaying computed matrix configurations:")
            preview_gen = itertools.islice(pipeline.get_cartesian_generator(), 50)

            for idx, combo in enumerate(preview_gen):
                full_seq, bit_contribs, vals = pipeline.extract_state_details(combo)
                print(f"\nCoord Index [{idx:04d}]")
                print(f"  -> Concatenated Binary : {full_seq}")
                for var in pipeline.variable_order:
                    print(f"  -> {var:<20}: Bits [{bit_contribs[var]}] | Decoded: {vals[var]}")

            if pipeline.total_matrix_size > 50:
                print(
                    f"\n... ({pipeline.total_matrix_size - 50:,} additional exhaustive combinations active in matrix space)")

        elif choice == "2":
            include_sequences = input(
                "Include full Programmatic State Binary Sequences in exports? (y/n): ").strip().lower() == 'y'

            # Verify and create the session-specific folder
            if not os.path.exists(pipeline.session_folder_name):
                os.makedirs(pipeline.session_folder_name)
                print(f"\n[INFO] Generated session archive directory: {pipeline.session_folder_name}")

            csv_filename = os.path.join(pipeline.session_folder_name, "cartesian_solution_traversals.csv")
            xlsx_filename = os.path.join(pipeline.session_folder_name, "cartesian_solution_traversals.xlsx")

            if not HAS_OPENPYXL:
                print(
                    "\n[WARNING] 'openpyxl' module not found. XLSX export will be skipped. Only CSV will be generated.")
                print("          (To enable XLSX, run 'pip install openpyxl')")

            try:
                csv_file = open(csv_filename, mode="w", newline="", encoding="utf-8")
                writer = csv.writer(csv_file)

                if HAS_OPENPYXL:
                    wb = Workbook(write_only=True)
                    sheet_idx = 1
                    ws = wb.create_sheet(title=f"Matrix_Part_{sheet_idx}")
                    excel_row_count = 1

                headers = ["Coordinate_Index", "Profile_Wattage"]
                if include_sequences:
                    headers.append("Total_Concatenated_Sequence")

                for var in pipeline.variable_order:
                    headers.extend([f"Decoded_Value_{var}", f"Binary_Contribution_{var}", f"BitAllocation_{var}"])

                writer.writerow(headers)
                if HAS_OPENPYXL:
                    ws.append(headers)

                print(f"Streaming Cartesian matrix to disk... (Press Ctrl+C to abort if matrix is too large)")
                for idx, combo in enumerate(pipeline.get_cartesian_generator()):
                    full_seq, bit_contribs, vals = pipeline.extract_state_details(combo)

                    row = [idx, pipeline.wattage]
                    if include_sequences:
                        row.append(f'="{full_seq}"')

                    for var in pipeline.variable_order:
                        formatted_bin_contrib = f'="{bit_contribs[var]}"'
                        row.extend([vals[var], formatted_bin_contrib, pipeline.continuum_map[var].allocated_bits])

                    writer.writerow(row)

                    if HAS_OPENPYXL:
                        if excel_row_count >= 1048500:
                            sheet_idx += 1
                            ws = wb.create_sheet(title=f"Matrix_Part_{sheet_idx}")
                            ws.append(headers)
                            excel_row_count = 1
                        ws.append(row)
                        excel_row_count += 1

                    if idx % 100000 == 0 and idx > 0:
                        print(f"  ... Processed {idx:,} permutations ...")

                csv_file.close()

                if HAS_OPENPYXL:
                    print("  ... Compressing and finalizing XLSX file (this may take a minute) ...")
                    wb.save(xlsx_filename)

                print(f"\n[SUCCESS] Comprehensive exhaustive export mapped to session archive:")
                print(f"  -> CSV:  {os.path.abspath(csv_filename)}")
                if HAS_OPENPYXL:
                    print(f"  -> XLSX: {os.path.abspath(xlsx_filename)}")

            except KeyboardInterrupt:
                print(f"\n[ABORTED] Export halted by user. Partial files saved in {pipeline.session_folder_name}.")
            except Exception as e:
                print(f"\n[EXPORT FAULT] Failed to generate solution matrices: {e}")
            finally:
                if not csv_file.closed:
                    csv_file.close()

        elif choice == "3":
            bin_input = input("\nEnter concatenated binary sequence string to map: ").strip().replace(" ", "")

            print("\n=========================================================")
            print("RECONSTRUCTED METRIC SYSTEM MATRIX SNAPSHOT:")
            ptr = 0
            for var in pipeline.variable_order:
                continuum = pipeline.continuum_map[var]
                bit_slice = bin_input[ptr: ptr + continuum.allocated_bits]
                decoded_val = continuum.decode_binary(bit_slice)
                print(f"  Field Element: {var:<25} -> Bits: [{bit_slice}] | Decoded: {decoded_val}")
                ptr += continuum.allocated_bits
            print("=========================================================")

        elif choice == "4":
            print("\n=== [Functionality1] N-Gram Binary Sequence Set Combinatorics Engine ===")
            matrix_path = input("Enter path to Base Binary Sequence File (.csv or .xlsx): ").strip().strip('"')
            loaded_seqs = load_sequence_file(matrix_path)

            if not loaded_seqs:
                print("[!] Structural sequence file unreadable or absent of valid data. Aborting module.")
                continue

            print(f"-> Successfully loaded {len(loaded_seqs)} unique binary elements.")

            try:
                start_n = int(input("Enter starting programmatic state sequence length (n) (e.g., 2): ").strip())
                max_n = int(input("Enter maximum programmatic sequence length (e.g., 5): ").strip())
            except ValueError:
                print("[ERROR] Boundaries must be integer numeric mappings.")
                continue

            print("\nSet Generation Method Selection Mode:")
            print("1. Strict Permutations (Order matters, no repeating internal sequences)")
            print("2. Strict Combinations (Order doesn't matter, repetition allowed)")
            print("3. Hybrid Mode (Cartesian structural variance over length)")
            gc_choice = input("Choice (1/2/3): ").strip()
            gen_method = 'combinations' if gc_choice == '2' else (
                'permutations_strict' if gc_choice == '1' else 'hybrid')

            ngram_engine = BinarySequenceNgramEngine(loaded_seqs, start_n, max_n, gen_method)
            ngram_engine.generate_and_map()

            # Archive generated logic sets inside the session boundary
            if not os.path.exists(pipeline.session_folder_name):
                os.makedirs(pipeline.session_folder_name)

            ngram_export_path = os.path.join(pipeline.session_folder_name,
                                             f"ngram_combinatorics_{pipeline.session_epoch}.csv")
            with open(ngram_export_path, 'w', newline='', encoding='utf-8') as n_file:
                writer = csv.writer(n_file)
                writer.writerow(["Sequence_Length", "Programmatic_Set_String", "Tuple_Array"])
                for length, combinations in ngram_engine.structural_matrix.items():
                    for comb in combinations:
                        joined_str = "-".join(comb)
                        writer.writerow([length, joined_str, str(comb)])

            print(f"\n[SUCCESS] N-Gram mappings exported to session archive: {ngram_export_path}")

        elif choice == "5":
            print("\nShutting down Unified Core Pipeline. Execution scope flushed securely.")
            break
        else:
            print("[Invalid entry] Select a valid programmatic execution branch.")