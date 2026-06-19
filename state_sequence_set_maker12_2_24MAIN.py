# -*- coding: utf-8 -*-
"""
UNIVERSAL RAW BINARY-FIRST CARTESIAN PROGRAMMATIC ENGINE (V6.3)
- Exhaustively evaluates all 2^n permutations for allocated bit lengths.
- Prevents OOM (Out of Memory) crashes via lazy evaluation generators.
- Isolates individual variable binary contributions in previews and CSV exports.
- Dual-streams to CSV and XLSX. Auto-chunks XLSX sheets at 1,048,500 rows.
*Upgrade: Isolates output documentation into dynamic, epoch-stamped session folders.*
"""

import os
import sys
import types
import csv
import itertools
import math
import time

# Attempt to load openpyxl for Excel output, falling back gracefully if missing
try:
    from openpyxl import Workbook

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class BiDict(dict):
    """Native Bidirectional Dictionary tracking object namespace relationships."""

    def __init__(self, *args, **kwargs):
        super(BiDict, self).__init__(*args, **kwargs)
        self.inverse = {}
        for key, value in self.items():
            self.inverse[value] = key

    def __setitem__(self, key, value):
        super(BiDict, self).__setitem__(key, value)
        self.inverse[value] = key

    def get_key_by_value(self, value):
        return self.inverse.get(value)


class CartesianBitContinuum:
    """Handles exhaustive low-level dynamic memory permutations."""

    def __init__(self, name, var_type, base_value, allocated_bits):
        self.name = name
        self.type = var_type
        self.base_value = base_value
        self.allocated_bits = allocated_bits
        # Total permutations is exactly 2^n
        self.total_permutations = 1 << self.allocated_bits

    def decode_binary(self, bit_str):
        """Decodes pure Cartesian raw bit configurations into programmatic values."""
        if len(bit_str) < self.allocated_bits:
            return "ERR (Truncated)"

        if self.type == "NUMERIC":
            return int(bit_str, 2)
        else:
            chars = []
            for i in range(0, len(bit_str), 8):
                byte = bit_str[i:i + 8]
                if len(byte) < 8: break
                chars.append(chr(int(byte, 2)))

            # Use repr() to safely escape unprintable ASCII generated during brute force
            decoded_str = "".join(chars)
            return repr(decoded_str)


class CartesianCorePipeline:
    def __init__(self):
        # Capture the epoch timestamp at the exact moment of session invocation
        self.session_epoch = int(time.time())
        self.session_folder_name = f"Session_{self.session_epoch}"

        self.continuum_map = BiDict()
        self.staging_metadata = {}
        self.variable_order = []
        self.wattage = 65.0
        self.total_matrix_size = 0

    def execute_introspection(self, file_path):
        if not os.path.exists(file_path):
            print(f"[ERROR] Target path invalid: {file_path}")
            return False

        self.continuum_map.clear()
        self.staging_metadata.clear()
        self.variable_order.clear()

        print(f"\n[PHASE 1] Dynamic Introspection Pipeline executing at: {file_path}")

        dir_name = os.path.dirname(os.path.abspath(file_path))
        sys.path.insert(0, dir_name)
        original_cwd = os.getcwd()
        os.chdir(dir_name)

        try:
            namespace = {"__file__": os.path.abspath(file_path), "__name__": "__main__"}
            with open(file_path, "r", encoding="utf-8") as f:
                exec(f.read(), namespace)

            for obj_name, obj_instance in namespace.items():
                if obj_name.startswith("__") or isinstance(obj_instance, (types.ModuleType, types.FunctionType, type)):
                    continue

                if hasattr(obj_instance, "__dict__") and not isinstance(obj_instance, dict):
                    print(f"  [+] Introspecting Object Instance Structure: '{obj_name}'")
                    for attr_name, attr_val in obj_instance.__dict__.items():
                        if attr_name.startswith("__"): continue
                        composite_name = f"{obj_name}.{attr_name}"
                        self.variable_order.append(composite_name)

                        var_type = "NUMERIC" if isinstance(attr_val, (int, float)) and not isinstance(attr_val,
                                                                                                      bool) else "LITERAL"
                        self.staging_metadata[composite_name] = {"val": attr_val, "type": var_type}
                        print(f"    [Captured {var_type}] '{composite_name}': Base Value -> {attr_val}")
                else:
                    self.variable_order.append(obj_name)
                    var_type = "NUMERIC" if isinstance(obj_instance, (int, float)) and not isinstance(obj_instance,
                                                                                                      bool) else "LITERAL"
                    self.staging_metadata[obj_name] = {"val": obj_instance, "type": var_type}
                    print(f"    [Captured {var_type}] '{obj_name}': Base Value -> {obj_instance}")

            return len(self.variable_order) > 0
        except Exception as e:
            print(f"[ERROR] Introspection pipeline failure: {e}")
            return False
        finally:
            os.chdir(original_cwd)

    def configure_memory_footprints(self):
        print("\n[PHASE 2] Binary Sequence Parameter Configuration")
        print("Define the exact exhaustive binary boundaries for each captured variable:")

        for var in self.variable_order:
            meta = self.staging_metadata[var]
            default_bits = 4 if meta["type"] == "NUMERIC" else 16

            try:
                user_bits = input(
                    f"  -> Enter bit-length for '{var}' (Type: {meta['type']}, Default: {default_bits}): ").strip()
                allocated_bits = int(user_bits) if user_bits else default_bits
            except ValueError:
                allocated_bits = default_bits

            continuum_obj = CartesianBitContinuum(var, meta["type"], meta["val"], allocated_bits)
            self.continuum_map[var] = continuum_obj

        print("\n[PHASE 3] Computing exhaustive Cartesian logic states from raw bit-blueprints...")

        self.total_matrix_size = math.prod(self.continuum_map[v].total_permutations for v in self.variable_order)
        print(f"--> Success. Total Exhaustive Generated Binary Combinations: {self.total_matrix_size:,}")
        if self.total_matrix_size > 1048500:
            print(
                "  [!] MASSIVE MATRIX DETECTED: XLSX generation will auto-chunk across multiple sheets to bypass the 1,048,576 row limit.")

    def get_cartesian_generator(self):
        """Returns a lazy-evaluating generator of all permutations to prevent RAM exhaustion."""
        spaces = [range(self.continuum_map[v].total_permutations) for v in self.variable_order]
        return itertools.product(*spaces)

    def extract_state_details(self, combo_tuple):
        """Processes a raw coordinate tuple into binary strings and decoded values."""
        binary_contributions = {}
        decoded_values = {}
        full_seq = ""

        for idx, var in enumerate(self.variable_order):
            continuum = self.continuum_map[var]
            bin_str = format(combo_tuple[idx], f"0{continuum.allocated_bits}b")

            binary_contributions[var] = bin_str
            full_seq += bin_str
            decoded_values[var] = continuum.decode_binary(bin_str)

        return full_seq, binary_contributions, decoded_values


# ==========================================
# EXECUTION ORCHESTRATOR
# ==========================================
if __name__ == "__main__":
    pipeline = CartesianCorePipeline()

    print("=================================================================")
    print("   UNIVERSAL RAW BINARY-FIRST CARTESIAN PROGRAMMATIC ENGINE      ")
    print("=================================================================")

    try:
        watt_in = input("Enter hardware configuration wattage (e.g., 65.0): ").strip()
        pipeline.wattage = float(watt_in) if watt_in else 65.0
    except ValueError:
        pipeline.wattage = 65.0

    path_input = input("Enter target absolute script file path (absfilepath): ").strip().strip('"')

    if pipeline.execute_introspection(path_input):
        pipeline.configure_memory_footprints()

        while True:
            print("\n------------------------------------------------------------")
            print("Select State Synthesis Mode:")
            print("1) Preview Generated Binary Continuum Sequences & Decoded Value States")
            print("2) Export Solution Set Traversals with Resource Profiling to CSV & XLSX")
            print("3) Input Concatenated Binary Sequence -> Rebuild System State")
            print("4) Exit Unified Core Pipeline")
            choice = input("Selection (1-4): ").strip()

            if choice == "1":
                print("\n[PREVIEW MODE] Displaying computed matrix configurations:")
                preview_gen = itertools.islice(pipeline.get_cartesian_generator(), 50)

                for idx, combo in enumerate(preview_gen):
                    full_seq, bit_contribs, vals = pipeline.extract_state_details(combo)
                    print(f"\nCoord Index [{idx:04d}]")
                    print(f"  -> Concatenated Binary : {full_seq}")
                    for var in pipeline.variable_order:
                        print(f"  -> {var:<20}: Bits [{bit_contribs[var]}] | Decoded: {vals[var]}")

                if pipeline.total_matrix_size > 50:
                    print(
                        f"\n... ({pipeline.total_matrix_size - 50:,} additional exhaustive combinations active in matrix space)")

            elif choice == "2":
                include_sequences = input(
                    "Include full Programmatic State Binary Sequences in exports? (y/n): ").strip().lower() == 'y'

                # Verify and create the session-specific folder
                if not os.path.exists(pipeline.session_folder_name):
                    os.makedirs(pipeline.session_folder_name)
                    print(f"\n[INFO] Generated session archive directory: {pipeline.session_folder_name}")

                csv_filename = os.path.join(pipeline.session_folder_name, "cartesian_solution_traversals.csv")
                xlsx_filename = os.path.join(pipeline.session_folder_name, "cartesian_solution_traversals.xlsx")

                if not HAS_OPENPYXL:
                    print(
                        "\n[WARNING] 'openpyxl' module not found. XLSX export will be skipped. Only CSV will be generated.")
                    print("          (To enable XLSX, run 'pip install openpyxl')")

                try:
                    # Initialize CSV Setup
                    csv_file = open(csv_filename, mode="w", newline="", encoding="utf-8")
                    writer = csv.writer(csv_file)

                    # Initialize XLSX Setup (Write-Only Mode for massive files)
                    if HAS_OPENPYXL:
                        wb = Workbook(write_only=True)
                        sheet_idx = 1
                        ws = wb.create_sheet(title=f"Matrix_Part_{sheet_idx}")
                        excel_row_count = 1

                    # Build dynamic header structure
                    headers = ["Coordinate_Index", "Profile_Wattage"]
                    if include_sequences:
                        headers.append("Total_Concatenated_Sequence")

                    for var in pipeline.variable_order:
                        headers.extend([f"Decoded_Value_{var}", f"Binary_Contribution_{var}", f"BitAllocation_{var}"])

                    writer.writerow(headers)
                    if HAS_OPENPYXL:
                        ws.append(headers)

                    # Stream records lazily across Cartesian plane
                    print(f"Streaming Cartesian matrix to disk... (Press Ctrl+C to abort if matrix is too large)")
                    for idx, combo in enumerate(pipeline.get_cartesian_generator()):
                        full_seq, bit_contribs, vals = pipeline.extract_state_details(combo)

                        row = [idx, pipeline.wattage]
                        if include_sequences:
                            # Force strict text interpretation for accurate spreadsheet rendering
                            row.append(f'="{full_seq}"')

                        for var in pipeline.variable_order:
                            formatted_bin_contrib = f'="{bit_contribs[var]}"'
                            row.extend([vals[var], formatted_bin_contrib, pipeline.continuum_map[var].allocated_bits])

                        writer.writerow(row)

                        # Handle XLSX chunking and row appending
                        if HAS_OPENPYXL:
                            if excel_row_count >= 1048500:  # Excel limit is 1,048,576
                                sheet_idx += 1
                                ws = wb.create_sheet(title=f"Matrix_Part_{sheet_idx}")
                                ws.append(headers)
                                excel_row_count = 1
                            ws.append(row)
                            excel_row_count += 1

                        if idx % 100000 == 0 and idx > 0:
                            print(f"  ... Processed {idx:,} permutations ...")

                    csv_file.close()

                    if HAS_OPENPYXL:
                        print("  ... Compressing and finalizing XLSX file (this may take a minute) ...")
                        wb.save(xlsx_filename)

                    print(f"\n[SUCCESS] Comprehensive exhaustive export mapped to session archive:")
                    print(f"  -> CSV:  {os.path.abspath(csv_filename)}")
                    if HAS_OPENPYXL:
                        print(f"  -> XLSX: {os.path.abspath(xlsx_filename)}")

                except KeyboardInterrupt:
                    print(f"\n[ABORTED] Export halted by user. Partial files saved in {pipeline.session_folder_name}.")
                except Exception as e:
                    print(f"\n[EXPORT FAULT] Failed to generate solution matrices: {e}")
                finally:
                    if not csv_file.closed:
                        csv_file.close()

            elif choice == "3":
                bin_input = input("\nEnter concatenated binary sequence string to map: ").strip().replace(" ", "")

                print("\n=========================================================")
                print("RECONSTRUCTED METRIC SYSTEM MATRIX SNAPSHOT:")
                ptr = 0
                for var in pipeline.variable_order:
                    continuum = pipeline.continuum_map[var]
                    bit_slice = bin_input[ptr: ptr + continuum.allocated_bits]
                    decoded_val = continuum.decode_binary(bit_slice)
                    print(f"  Field Element: {var:<25} -> Bits: [{bit_slice}] | Decoded: {decoded_val}")
                    ptr += continuum.allocated_bits
                print("=========================================================")

            elif choice == "4":
                print("\nShutting down Unified Core Pipeline. Execution scope flushed securely.")
                break
            else:
                print("[Invalid entry] Select a valid programmatic execution branch.")